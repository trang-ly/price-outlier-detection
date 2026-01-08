"""
Price Matrix Outlier Detection
---------------------------------
Detects pricing anomalies across store/location columns using
median-based clustering with a tolerance threshold.

Highlights:
- Outliers: red font
- #### cells: yellow fill

Usage:
python price_matrix_outlier_detection.py \
--input sample_price_matrix.xlsx \
--output processed_output.xlsx \
--tolerance 1.00 \
--skip-prefixes cleaning,service
"""

from __future__ import annotations

import argparse
import warnings
from pathlib import Path
from typing import Dict, List, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# ==================================
# CONFIGURATION
# ==================================
sanity_pct = 0.20  # cluster median must be >= 20% of row median

red_font = Font(color="FF0000")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Suppress harmless legacy Excel warnings
warnings.filterwarnings("ignore", message="OLE2 inconsistency")

# ==================================
# ARGUMENT PARSING
# ==================================


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Detect price outliers in a price matrix Excel file")
    parser.add_argument("--input", required=True, help="Input Excel file (.xls or .xlsx)")
    parser.add_argument("--output", default="processed_output.xlsx", help="Output Excel file")
    parser.add_argument("--tolerance", type=float, default=1.00, help="Cluster tolerance (e.g. 0.75)")
    parser.add_argument("--skip-prefixes", default="",
                        help="Comma-separated list of prefixes to skip, e.g., 'cleaning,service'")
    return parser.parse_args()

# ==================================
# DATA LOADING
# ==================================


def load_excel(path: Path) -> Tuple[pd.DataFrame, List[str]]:
    """
    Load Excel file into a DataFrame and determine value_cols.
    Assumes first column is InventoryItemName and all remaining columns are value_cols.
    Returns the DataFrame and a list of value_cols.
    """
    df = pd.read_excel(path)
    value_cols = list(df.columns[1:])  # all columns except the first
    return df.copy(), value_cols


# ==================================
# CLUSTERING LOGIC
# ==================================


def build_clusters(values: List[float], tol: float) -> List[List[float]]:
    """
    Groups numeric values into clusters where adjacent values differ by no more than +/-tol.
    """
    if not values:
        return []

    vals = sorted(values)
    clusters = [[vals[0]]]

    for v in vals[1:]:
        # Compare against the last value in the cluster (single-linkage)
        if abs(v - clusters[-1][-1]) <= tol:
            clusters[-1].append(v)
        else:
            clusters.append([v])
    return clusters


def merge_nearby_clusters(clusters: List[List[float]], pct: float) -> List[List[float]]:
    """
    Merge clusters whose medians are within pct of each other.
    """
    if not clusters:
        return []

    merged = [clusters[0]]
    for c in clusters[1:]:
        prev_median = np.median(merged[-1])
        curr_median = np.median(c)
        # Collapse price bands with similar medians using relative threshold
        if abs(curr_median - prev_median) <= prev_median * pct:
            merged[-1].extend(c)
        else:
            merged.append(c)
    return merged


# ==================================
# OUTLIER DETECTION
# ==================================


def detect_outliers(
        df: pd.DataFrame,
        value_cols: List[str],
        tolerance: float,
        skip_prefixes: List[str]) -> \
        Tuple[Dict[Tuple[int, str], bool], Dict[Tuple[int, str], bool], pd.Series]:
    """
    Detect pricing outliers across location columns on a per-item basis.

    Logic:
      - Build price clusters using adaptive, row-level tolerance
      - Discard implausibly low-value clusters
      - Select the most representative price cluster
      - Flag values outside that cluster as outliers
    """
    outliers: Dict[Tuple[int, str], bool] = {}
    hash_cells: Dict[Tuple[int, str], bool] = {}
    row_clusters = pd.Series(index=df.index, dtype=object)  # store clusters per row

    # Percentage threshold for merging nearby clusters
    merge_pct = 0.15

    for idx, row in df.iterrows():
        item_name = str(row.get("InventoryItemName", "")).strip().lower()
        if any(item_name.startswith(p.lower()) for p in skip_prefixes):
            continue

        numeric_vals: List[float] = []
        col_value_map: Dict[str, float] = {}

        for col in value_cols:
            raw = row.get(col)
            if pd.isna(raw) or str(raw).strip() == "":
                continue
            if isinstance(raw, str) and "####" in raw:
                hash_cells[(idx, col)] = True
                continue
            try:
                num = float(str(raw).replace("$", "").replace(",", ""))
            except ValueError:
                continue
            if num == 0:
                continue
            numeric_vals.append(num)
            col_value_map[col] = num

        # Insufficient data to infer a price pattern
        if len(numeric_vals) <= 2:
            row_clusters[idx] = str(numeric_vals)
            continue

        row_median = np.median(numeric_vals)
        # Discard clusters far below expected price range
        sanity_cutoff = row_median * sanity_pct
        # Adaptive tolerance scales with price magnitude
        adaptive_tol = max(tolerance, row_median * 0.15)

        # Build and stabilize clusters
        clusters = build_clusters(numeric_vals, adaptive_tol)
        clusters = merge_nearby_clusters(clusters, merge_pct)

        valid_clusters = [c for c in clusters if np.median(c) >= sanity_cutoff]
        row_clusters[idx] = str(clusters)  # store all clusters before picking normal

        if not valid_clusters:
            continue

        # Select the largest cluster as the expected price range
        valid_clusters.sort(key=len, reverse=True)
        max_len = len(valid_clusters[0])
        largest_clusters = [c for c in valid_clusters if len(c) == max_len]

        # Tie-break using proximity to row median
        if len(largest_clusters) > 1:
            largest_clusters.sort(key=lambda c: abs(np.median(c) - row_median))

        correct_cluster = largest_clusters[0]

        for col, num in col_value_map.items():
            if num not in correct_cluster:
                outliers[(idx, col)] = True

    return outliers, hash_cells, row_clusters


# ==================================
# OUTPUT FORMATTING
# ==================================


def write_output(df: pd.DataFrame,
                 output_path: Path,
                 outliers: Dict[Tuple[int, str], bool],
                 hash_cells: Dict[Tuple[int, str], bool]) -> None:
    df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active

    header_map = {
        ws.cell(row=1, column=i).value: i
        for i in range(1, ws.max_column + 1)
    }

    for (r, c) in outliers:
        ws.cell(row=r + 2, column=header_map[c]).font = red_font

    for (r, c) in hash_cells:
        ws.cell(row=r + 2, column=header_map[c]).fill = yellow_fill

    wb.save(output_path)


# ==================================
# MAIN ENTRY POINT
# ==================================


def main() -> None:
    args = parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    tolerance = round(args.tolerance, 2)
    skip_prefixes = [p.strip() for p in args.skip_prefixes.split(",") if p.strip()]

    df, value_cols = load_excel(input_path)
    outliers, hash_cells, row_clusters = detect_outliers(df, value_cols, tolerance, skip_prefixes)

    # Append cluster breakdown per row for review (optional)
    # df["Clusters"] = row_clusters

    write_output(df, output_path, outliers, hash_cells)
    print(f"Done — output saved to: {output_path}")


if __name__ == "__main__":
    main()
